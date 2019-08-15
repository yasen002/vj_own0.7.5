import xlrd
import pymongo
import sys
from collections import Counter
import pathlib
import os
import os.path
import openpyxl
import xlsxwriter

newCatolog = [{'038000188633': {'Tuna': [10]}}, {'3605970360757': {'salmon avocado': [15]}}, {'{071662000240': {'albacore': []}}, {'681131287326': {'cucumber': [12]}}, {'856293003255': {'salad': [3]}}, {'9780375501203': {'avocado': [9]}}, {'639785305286': {'california': [3]}}, {'8872878': {'tunami': []}}, {'4243107253': {'fish': []}}, {'9780070527096': {'lettuce': [4]}}, {'15363636': {'ginger': [1]}}, {'00020190': {'Wasabi': [1]}}, {'681131119740': {'crab': [9]}}, {'9780070527096.0': {'shrimp': []}}, {'11742166.0': {'knife': []}}, {'639785305286.0 ': {'rice': []}}, {'total': {'total': []}}]

catolog = {'038000188633':'tuna avocado',
           '3605970360757': 'salmon avocado',
           '071662000240': 'rainbow roll',
           '681131287326': 'salad roll',
           '856293003255': 'Nigiri',
           '9780375501203': 'vege roll',
           '639785305286': 'cucumber roll',
           '8872878':'home',
           '4243107253':'work',
           '9780070527096': 'fish',
           '15363636':'tree',
           '00020190':'egg',
           '681131119740':'been',
           '9780070527096.0':'salad',
           '11742166.0':'drink',
           '639785305286.0 ':'fu',
           'total': 'total'
           
           }


catolog2 = [
  {"038000188633":{"Tuna":[]}  },
  {'3605970360757': {'salmon avocado':[]}  },
  {'{071662000240': {"albacore":[]}  },
  {'681131287326': {"cucumber":[]}  },
  {'856293003255': {"salad":[]} },
  {'9780375501203': {"avocado":[]} },
  {'639785305286': {"california":[]} },
  {'8872878':{"tunami":[]} },
  {'4243107253':{"fish":[]} },
  {'9780070527096': {"lettuce":[]} },
  {'15363636':{"ginger":[]} },
  {'00020190':{"Wasabi":[]} },
  {'681131119740':{"crab":[]} },
  {'9780070527096.0':{"shrimp":[]} },
  {'11742166.0':{"knife":[]} },
  {'639785305286.0 ':{"rice":[]} },
  {'total': {"total":[]} }
]


def update(xlsxResultFileName,oneDayListSum):
  if not os.path.isfile(xlsxResultFileName):
  # -----copy catolog items to new file
    print('creating new file')
  # open workbook
    workbook = xlsxwriter.Workbook(xlsxResultFileName)
    worksheet = workbook.add_worksheet()
    row = 0
    col = 0

    # catolog item name to a list
    CatologList = []
    for code,item in (catolog.items()):
      CatologList.append(item)

    # write item to the new workbook
    for i in range(len(CatologList)):
      worksheet.write(row, col, CatologList[i])
      row += 1
    
    worksheet.write(row, col, 'Total')

    workbook.close()
    update(xlsxResultFileName,oneDayListSum) 

  else:
    print('updating existing file')
     # catolog item name to a list
    CatologList = []
    for code,item in (catolog.items()):
      CatologList.append(item)


    # -------get the item coordinate and quant
    x_quan_co = {}
    for i,v in oneDayListSum.items():
        if i in CatologList:
          x = CatologList.index(i)
          ItemQuant = v
          x_quan_co.update({x:ItemQuant})

        

    # -------filling data with coordinate

    # finding y coodinate
    def nrowsCals(fileName):
        wb = xlrd.open_workbook(fileName)
        ws = wb.sheet_by_index(0)
        rows = ws.nrows
        cols = ws.ncols
        return(rows,cols)

    a = nrowsCals(xlsxResultFileName)
    rows = a[0]
    cols = a[1] + 1

    # updating with openpyxl
    wb = openpyxl.load_workbook(xlsxResultFileName)
    ws = wb['Sheet1']
    # fill in quant in the coordinate
    for x,q in (x_quan_co.items()):
      ws.cell(x+1, cols).value = q

    # # Write a total using a formula.
    ws.cell(rows, cols).value = sum(x_quan_co.values())
    wb.save(xlsxResultFileName)




def reading(incomingExelData):
  #--------------- open incomming exel file
  inputWorkbook = xlrd.open_workbook(incomingExelData)
  inputWorksheet = inputWorkbook.sheet_by_index(0)

  #-----------extract the barcode in to a list
  itemList = []
  zList =[]
  for y in range(1,inputWorksheet.nrows):
    itemList.append(inputWorksheet.cell_value(y,0))
  
  #------------count and sort the list in key value pair
  counted = dict(Counter(itemList))
  dataResult = counted.copy()

  for i in range(len(catolog2)):
    a = catolog2[i]
    # i = bcode in zLog
    for ii in a:
      bob = dataResult.get(ii)
      if bob:
        # rollNum list add b
        b = a[ii]
        for i in b:
          sa = b[i]
          sa.append(bob)
      b = a[ii]
      zList.append(b) 
  return zList


newCatolog = reading('b.xlsx')
print(newCatolog)


def newUpdate(newCatolog):
  if(newCatolog):
    print('need to add')

  # else:
    # create newCatolog











# each object
# for key in catolog2:
  # print(key)

## value & key
# for i in range(len(catolog2)):
#   a = catolog2[i]
#   for i in a:
#     print(i)


# number of rolls
# for i in range(len(catolog2)):
#   a = catolog2[i]
#   for i in a:
#     b = a[i]
#     for i in b:
#       print(b[i])

    















    #variables
# xlsxResultFileName = "Expenses88.xlsx"
# dailyCollectionFolder = "upload"

# # loop through all daily files
# in_files = os.listdir(dailyCollectionFolder)
# reList = []
# for i in in_files:
#   # reading
#   oneDayListSum = reading(dailyCollectionFolder+"\\"+i)
#   reList.append(oneDayListSum)
#   # updating to xlsx file
#   update(xlsxResultFileName,oneDayListSum)

# # insert to mongoose with insert_many


# kikkaClient = pymongo.MongoClient("mongodb://localhost:27017/")
# kikkadb = kikkaClient["kikkadb"]
# mycol = kikkadb["sushi"]
# x = mycol.insert_many(reList)